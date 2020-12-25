# import pandas as pd
# import codecs
# xd = pd.ExcelFile('mdemo新物料/1.xlsx')
# df = xd.parse()
# with codecs.open('mdemo新物料/1.html','w','utf-8') as html_file:
#     html_file.write(df.to_html(header = True,index = False))

import openpyxl


def parse_merged_cells():
    # 整合所有合并单元
    cells = []
    for i in excel.merged_cell_ranges:
        for row in list(i.rows):
            cells+=row
    return cells

def judge(cell):
    # 判断输入的单元格是否是合并单元格
    # 0: 不是单元格，或是合并单元格的第一位
    # 1：同行，列+1
    # 2：同列，行+1
    # 3：既不同行也不同列，不处理
    if (cell.row,cell.column) not in merged_cells:
        return 0,cell.value
    for i in excel.merged_cells:
        for row in list(i.rows):
            if (cell.row,cell.column) not in row:
                continue
            if (cell.row,cell.column) == i.top[0]:
                return 0,i.top[0]
            elif cell.row == i.top[0][0]:
                return 1,i.top[0]
            elif cell.column == i.top[0][1]:
                return 2,i.top[0]
            else:
                return 3,i.top[0]
            # if i.size['rows'] == 1:
            #     if (cell.row,cell.column) == i.left[0]:
            #         return 0,cell.value
            #     else:
            #         # return 1,excel.cell(row=i.left[0][0],column=i.left[0][1]).value
            #         return 1,i.left[0]
            
            # elif i.size['columns']==1:
            #     if (cell.row,cell.column) == i.top[0]:
            #         return 0,cell.value
            #     else:
            #         # return 2,excel.cell(row=i.top[0][0],column=i.top[0][1]).value
            #         return 2,i.top[0]

            # else:# 行、列均不为1 即：合并格为二维的
            #     if (cell.row,cell.column) == i.top[0]:
            #         return 0,cell.value
            #     else:
            #         return 2,i.top[0]

i_file = 'excel/1.xlsx'
o_file = 'excel/1.html'
excel = openpyxl.load_workbook(i_file).active
# max_col = excel.max_column
# max_row = excel.max_row

merged_cells = parse_merged_cells()
print(merged_cells)
'''
[
    [{'value':[num,num]},{},{}],
    [{'value':[num,num]},{},{}],
    [{'value':[num,num]},{},{}],
]
'''
data = []
for r in excel.iter_rows():
    row = []
    for cell in r:
        merge,position = judge(cell)
        if merge == 0:
            row.append({cell.value:[1,1]})
        if merge == 1:
            txt = excel.cell(row=position[0],column=position[1]).value
            row[-1][txt][0] += 1
        if merge == 2:
            txt = excel.cell(row=position[0],column=position[1]).value
            column = 0
            for index in range(len(data[position[0]-1])):
                for key in data[position[0]-1][index]:
                    column += data[position[0]-1][index][key][0]
                if column >= position[1]:
                    data[position[0]-1][index][txt][1]+=1
                    break
    if row:
        data.append(row)
print(data)

table = ''
for index,row in enumerate(data):
    if index == 0:
        table += "<tr class='thead'>"
    else:
        table += '<tr>'
    for cell in row:
        for key in cell.keys():
            table += "<td"
            # print(cell[key])
            if cell[key][0]>1:
                table += " colspan='%s'"%(cell[key][0])
            if cell[key][1]>1:
                table += " rowspan='%s'"%(cell[key][1])
            table+='>'
            table = table+str(key)+'</td>'
    table += "</tr>"
table = "<div class='table'><table><tbody>"+table+"</tbody></table></div>"
with open(o_file,'w',encoding='utf-8')as f:
    f.write(table)